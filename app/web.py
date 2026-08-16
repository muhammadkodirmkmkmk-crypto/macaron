"""FastAPI: JSON-API + дашборд + выгрузка в Excel."""
from __future__ import annotations

import datetime as dt
import hashlib
import io

from fastapi import Body, FastAPI, HTTPException, Query, Request, Response
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
from sqlalchemy import select

from . import analytics, config, faults as faults_mod
from .db import Batch, Fault, session

STATIC = config.BASE_DIR / "static"
COOKIE = "makaron_auth"
# доступно без пароля: логотип нужен самой странице входа
PUBLIC_PATHS = {"/health", "/login", "/logo.png", "/logo-64.png"}


def _token() -> str:
    return hashlib.sha256(("makaron::" + config.DASHBOARD_PASSWORD).encode()).hexdigest()[:32]


def _authed(request: Request) -> bool:
    if not config.DASHBOARD_PASSWORD:
        return True
    if request.cookies.get(COOKIE) == _token():
        return True
    if request.headers.get("X-Auth") == config.DASHBOARD_PASSWORD:
        return True
    if request.query_params.get("key") == config.DASHBOARD_PASSWORD:
        return True
    return False


def create_app() -> FastAPI:
    app = FastAPI(title="Sana Bogatir — аналитика сушки", docs_url=None, redoc_url=None)

    LOGIN_HTML = """<!doctype html>
<html lang=uz><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1,viewport-fit=cover">
<title>Sana Bogatir — hisobotlar paneli</title>
<link rel="icon" href="/logo-64.png">
<style>
*{box-sizing:border-box}
:root{--surface:#fcfcfb;--plane:#f4f5f7;--ink:#0A273D;--ink2:#52627a;--line:#c3c2b7;
      --border:rgba(10,39,61,.12);--brand:#0A273D;--gold:#FBC443;--bad:#d03b3b}
@media(prefers-color-scheme:dark){:root{--surface:#12293c;--plane:#0A273D;--ink:#fff;--ink2:#b9c6d4;
      --line:#2a4256;--border:rgba(255,255,255,.12);--brand:#FBC443;--gold:#FBC443;--bad:#e66767}}
html,body{height:100%}
body{margin:0;background:var(--plane);color:var(--ink);
  font:16px/1.45 system-ui,-apple-system,"Segoe UI",sans-serif;-webkit-font-smoothing:antialiased;
  display:flex;align-items:center;justify-content:center;
  padding:24px max(20px,env(safe-area-inset-left)) calc(24px + env(safe-area-inset-bottom))}
form{background:var(--surface);border:1px solid var(--border);border-radius:20px;
  box-shadow:0 8px 30px rgba(0,0,0,.07);width:100%;max-width:420px;padding:32px 28px}
.logo{width:76px;height:76px;display:block;margin:0 auto 16px;border-radius:50%}
.center{text-align:center}
h1{font-size:23px;font-weight:680;letter-spacing:-.02em;margin:0 0 6px}
p{color:var(--ink2);font-size:15px;margin:0 0 22px}
label{display:block;font-size:13px;color:var(--ink2);margin-bottom:7px;font-weight:500}
input{width:100%;height:54px;padding:0 16px;border:1.5px solid var(--line);border-radius:12px;
  font-size:17px;background:var(--plane);color:var(--ink);outline:none}
input:focus{border-color:var(--brand);box-shadow:0 0 0 3px color-mix(in srgb,var(--brand) 22%,transparent)}
button{width:100%;margin-top:16px;height:54px;border:0;border-radius:12px;background:var(--brand);
  color:#fff;font-size:17px;font-weight:650;cursor:pointer;-webkit-tap-highlight-color:transparent}
@media(prefers-color-scheme:dark){button{color:#0A273D}}
button:active{opacity:.85}
.err{background:color-mix(in srgb,var(--bad) 12%,transparent);color:var(--bad);
  border-radius:10px;padding:11px 14px;font-size:14px;margin:0 0 18px;font-weight:500}
@media(min-width:760px){
  form{max-width:380px;padding:34px 30px}
  .logo{width:66px;height:66px}
  h1{font-size:21px}
  input,button{height:48px;font-size:16px}
}
</style>
<form method=get action=/login>
  <img class=logo src="/logo.png" alt="Sana Bogatir">
  <div class=center>
    <h1>Sana Bogatir</h1>
    <p>Quritish sexi &middot; hisobotlar paneli</p>
  </div>
  __ERROR__
  <label for=p>Parol</label>
  <input id=p name=p type=password autofocus autocomplete=current-password
         inputmode=text placeholder="&bull;&bull;&bull;&bull;">
  <button>Kirish</button>
</form>"""

    def login_page(error: bool = False) -> str:
        block = '<p class="err">Parol to\'g\'ri kelmadi. Yana urinib ko\'ring.</p>' if error else ""
        return LOGIN_HTML.replace("__ERROR__", block)

    @app.get("/health")
    async def health():
        """Живость + действующие настройки. Секретов здесь нет — только режимы,
        чтобы можно было проверить, включено ли напоминание, не заходя в Railway."""
        return {
            "ok": True,
            "time": dt.datetime.now(config.TZ).isoformat(),
            "timezone": config.TIMEZONE,
            "dryers": config.DRYER_COUNT,
            "boilers": [f"{lo}-{hi}" for _, lo, hi in config.BOILER_RANGES],
            "reminder": {
                "enabled": config.LOAD_REMINDER_ENABLED,
                "hours": config.LOAD_REMINDER_HOURS,
            },
            "daily_report_hour": config.DAILY_REPORT_HOUR,
            "norm": {
                "min_minutes": config.NORM_MIN_MINUTES,
                "max_minutes": config.NORM_MAX_MINUTES,
                "cycle_minutes": config.NORM_CYCLE_MINUTES,
            },
            "stale_hours": config.STALE_HOURS,
            "vision": {"enabled": config.VISION_ENABLED, "model": config.ANTHROPIC_MODEL},
            "assistant": {"enabled": config.ASSISTANT_ENABLED, "model": config.ASSISTANT_MODEL},
            "telegram_bot": bool(config.TELEGRAM_BOT_TOKEN),
        }

    @app.get("/login")
    async def login(p: str = ""):
        if not config.DASHBOARD_PASSWORD:
            return Response(status_code=302, headers={"Location": "/"})
        if p == config.DASHBOARD_PASSWORD:
            r = Response(status_code=302, headers={"Location": "/"})
            r.set_cookie(COOKIE, _token(), max_age=60 * 60 * 24 * 90, httponly=True, samesite="lax")
            return r
        return HTMLResponse(login_page(error=bool(p)), status_code=401 if p else 200)

    @app.middleware("http")
    async def guard(request: Request, call_next):
        if request.url.path in PUBLIC_PATHS or _authed(request):
            resp = await call_next(request)
            if request.query_params.get("key") == config.DASHBOARD_PASSWORD and config.DASHBOARD_PASSWORD:
                resp.set_cookie(COOKIE, _token(), max_age=60 * 60 * 24 * 90, httponly=True, samesite="lax")
            return resp
        if request.url.path.startswith("/api/"):
            return JSONResponse({"error": "unauthorized"}, status_code=401)
        return HTMLResponse(login_page(), status_code=401)

    @app.get("/")
    async def index():
        return FileResponse(STATIC / "dashboard.html")

    @app.get("/logo.png")
    async def logo():
        return FileResponse(STATIC / "logo.png", headers={"Cache-Control": "public, max-age=604800"})

    @app.get("/logo-64.png")
    async def logo_small():
        return FileResponse(STATIC / "logo-64.png", headers={"Cache-Control": "public, max-age=604800"})

    @app.get("/api/dashboard")
    async def api_dashboard(days: int = Query(30, ge=1, le=365)):
        async with session() as s:
            data = await analytics.dashboard_payload(s, days=days)
            data["faults"] = await faults_mod.payload(s, days=days)
            return data

    # ---------------- поломки в цехе ----------------
    @app.get("/api/faults")
    async def api_faults(days: int = Query(30, ge=1, le=365)):
        async with session() as s:
            return await faults_mod.payload(s, days=days)

    @app.post("/api/faults")
    async def api_fault_add(payload: dict = Body(...)):
        part = (payload.get("part") or "").strip()
        text = (payload.get("text") or "").strip()
        if not part:
            raise HTTPException(400, "part kerak")
        async with session() as s:
            row = await faults_mod.add(s, part=part, text=text,
                                       who=(payload.get("who") or "Dashboard"), source="web")
            await s.commit()
            return faults_mod.serialize(row)

    @app.patch("/api/faults/{fault_id}")
    async def api_fault_patch(fault_id: int, payload: dict = Body(...)):
        async with session() as s:
            row = await s.get(Fault, fault_id)
            if not row:
                raise HTTPException(404, "topilmadi")
            if "status" in payload:
                st = payload["status"]
                if st == "fixed" and row.status != "fixed":
                    row.status = "fixed"
                    row.fixed_at = faults_mod.utcnow()
                    row.fixed_by = (payload.get("fixed_by") or "Dashboard")[:128]
                elif st == "open":
                    row.status, row.fixed_at, row.fixed_by = "open", None, None
            if payload.get("cost") is not None:
                row.cost = max(0, int(payload.get("cost") or 0))
            if payload.get("text"):
                row.text = str(payload["text"])[:500]
            await s.commit()
            return faults_mod.serialize(row)

    @app.delete("/api/faults/{fault_id}")
    async def api_fault_delete(fault_id: int):
        async with session() as s:
            row = await s.get(Fault, fault_id)
            if not row:
                raise HTTPException(404, "topilmadi")
            await s.delete(row)
            await s.commit()
            return {"ok": True}

    @app.get("/api/batches")
    async def api_batches(
        days: int = Query(30, ge=1, le=365),
        dryer: int | None = None,
        product: str | None = None,
        limit: int = Query(500, ge=1, le=5000),
    ):
        async with session() as s:
            rows = await analytics.load(s, days=days, product=product, dryer=dryer)
        return {"count": len(rows), "items": [analytics.serialize(b) for b in rows[:limit]]}

    @app.patch("/api/batches/{batch_id}")
    async def api_patch(batch_id: int, payload: dict = Body(...)):
        allowed = {"dryer_number", "product", "duration_minutes", "quality",
                   "note", "temperature", "humidity"}
        async with session() as s:
            b = await s.get(Batch, batch_id)
            if not b:
                raise HTTPException(404, "not found")
            for k, v in payload.items():
                if k in allowed:
                    setattr(b, k, v)
            if b.dryer_number and b.duration_minutes:
                b.needs_review = False
                b.started_at = b.finished_at - dt.timedelta(minutes=b.duration_minutes)
            b.source = "manual"
            await s.commit()
            return analytics.serialize(b)

    @app.delete("/api/batches/{batch_id}")
    async def api_delete(batch_id: int):
        async with session() as s:
            b = await s.get(Batch, batch_id)
            if not b:
                raise HTTPException(404, "not found")
            await s.delete(b)
            await s.commit()
        return {"ok": True}

    @app.post("/api/reset")
    async def api_reset(confirm: str = Query("", description="должно быть TOZALASH")):
        """Стереть все накопленные данные и начать с чистого листа.

        Схема, настройки и логика остаются — удаляются только записи:
        партии, сырые сообщения, открытые заходы и незакрытые вопросы бота.
        Требует явного подтверждения, чтобы не сработать по случайному запросу.
        """
        if confirm != "TOZALASH":
            raise HTTPException(400, "нужен параметр confirm=TOZALASH")
        from sqlalchemy import delete as sql_delete

        from .db import LoadEvent, Pending, RawMessage, StopEvent

        wiped = {}
        async with session() as s:
            # порядок важен: сначала таблицы, ссылающиеся на партии
            for model in (Pending, RawMessage, LoadEvent, StopEvent, Batch):
                res = await s.execute(sql_delete(model))
                wiped[model.__tablename__] = res.rowcount or 0
            await s.commit()
        return {"ok": True, "deleted": wiped}

    @app.post("/api/remap-display")
    async def api_remap():
        """Пересчитать температуру и влажность из сохранённых строк табло.
        Идемпотентно: запускать можно сколько угодно раз."""
        from .parser import map_display

        changed = []
        async with session() as s:
            rows = (await s.execute(
                select(Batch).where(Batch.display_raw.is_not(None))
            )).scalars().all()
            for b in rows:
                parts = (b.display_raw or "").split("|")
                parts += [""] * (3 - len(parts))
                timer, temp, hum = map_display(*(x or None for x in parts[:3]))
                if (temp, hum) != (b.temperature, b.humidity):
                    changed.append({
                        "id": b.id, "dryer": b.dryer_number, "raw": b.display_raw,
                        "temperature": [b.temperature, temp], "humidity": [b.humidity, hum],
                    })
                    b.temperature, b.humidity = temp, hum
                    if timer and not b.timer_raw:
                        b.timer_raw = timer
            await s.commit()
        return {"scanned": len(rows), "updated": len(changed), "changes": changed[:100]}

    @app.get("/api/export.xlsx")
    async def api_export(days: int = Query(30, ge=1, le=365)):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        async with session() as s:
            batches = await analytics.load(s, days=days)
            payload = await analytics.dashboard_payload(s, days=days)

        wb = Workbook()
        wb.properties.creator = "Sana Bogatir"
        wb.properties.title = f"Sana Bogatir — quritish sexi, {days} kun"
        head_font = Font(bold=True, color="FFFFFF")
        head_fill = PatternFill("solid", fgColor="0A273D")

        def style_head(ws, ncols):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = head_font
                cell.fill = head_fill
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"

        ws = wb.active
        ws.title = "Partiyalar"
        cols = ["ID", "Sushka", "Qozon", "Mahsulot", "Vaqt (min)", "Vaqt (soat)", "Boshlandi",
                "Tugadi", "t°", "Namlik", "Sifat", "Izoh", "Operator", "Ishonch", "Xabar matni"]
        ws.append(cols)
        for b in batches:
            ws.append([
                b.id, b.dryer_number, config.boiler_of(b.dryer_number),
                b.product, b.duration_minutes,
                round(b.duration_minutes / 60, 2) if b.duration_minutes else None,
                analytics.to_local(b.started_at).replace(tzinfo=None) if b.started_at else None,
                analytics.to_local(b.finished_at).replace(tzinfo=None),
                b.temperature, b.humidity, b.quality, b.note, b.user_name,
                b.confidence, b.raw_text,
            ])
        style_head(ws, len(cols))
        for i, w in enumerate([6, 8, 8, 14, 11, 11, 18, 18, 8, 9, 10, 26, 18, 9, 40], 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        ws2 = wb.create_sheet("Sushkalar")
        ws2.append(["Sushka", "Qozon", "Partiyalar", "O'rtacha (min)", "Mediana", "Min", "Maks",
                    "Global farq", "Brak", "Brak %", "Holat", "Oxirgi mahsulot", "Oxirgi tugash"])
        for d in payload["dryers"]:
            ws2.append([d["number"], d["boiler"], d["batches"], d["avg"], d["median"],
                        d["min"], d["max"], d["vs_global"], d["defects"], d["defect_rate"],
                        d["status"], d["last_product"], d["last_finished_at"]])
        style_head(ws2, 13)

        wsb = wb.create_sheet("Qozonlar")
        wsb.append(["Qozon", "Sushkalar", "Ishlayapti", "Partiyalar", "O'rtacha (min)",
                    "Mediana", "Min", "Maks", "Jami (min)", "Brak", "Brak %"])
        for b in payload["boilers"]:
            wsb.append([f"{b['id']} (№{b['from']}–{b['to']})", b["dryers"], b["in_work"],
                        b["batches"], b["avg"], b["median"], b["min"], b["max"],
                        b["total_minutes"], b["defects"], b["defect_rate"]])
        style_head(wsb, 11)

        ws3 = wb.create_sheet("Mahsulotlar")
        ws3.append(["Mahsulot", "Partiyalar", "O'rtacha", "Mediana", "Min", "Maks", "Std", "Brak", "Brak %", "Sushkalar"])
        for p in payload["products"]:
            ws3.append([p["product"], p["count"], p["avg"], p["median"], p["min"], p["max"],
                        p["stdev"], p["defects"], p["defect_rate"], p["dryers"]])
        style_head(ws3, 10)

        ws4 = wb.create_sheet("Kunlar")
        ws4.append(["Sana", "Partiyalar", "O'rtacha vaqt", "Brak"])
        for t in payload["timeline"]:
            ws4.append([t["date"], t["count"], t["avg"], t["defects"]])
        style_head(ws4, 4)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"sana_bogatir_{dt.datetime.now(config.TZ):%Y%m%d}_{days}d.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    return app
